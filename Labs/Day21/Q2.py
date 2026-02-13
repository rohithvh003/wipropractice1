import pandas as pd
from openpyxl import load_workbook, Workbook


# PANDAS METHOD
df = pd.read_excel("sales_data.xlsx", sheet_name="2025")
df["Total"] = df["Quantity"] * df["Price"]
df.to_excel("sales_summary.xlsx", index=False)

print("Successfully created new file using pandas")


# OPENPYXL METHOD
wb = load_workbook("sales_data.xlsx")
sheet = wb["2025"]

new_wb = Workbook()
new_sheet = new_wb.active

new_sheet.append(["Product", "Quantity", "Price", "Total"])

for r in sheet.iter_rows(min_row=2, values_only=True):
    if r[1] is not None and r[2] is not None:
        total = r[1] * r[2]
        new_sheet.append([r[0], r[1], r[2], total])

new_wb.save("sales_summary_openpyxl.xlsx")

print("OpenPyXL file created")
